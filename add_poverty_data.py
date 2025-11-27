#!/usr/bin/env python3
"""
Script to add World Bank poverty indicators to the existing GDP data Excel file.
"""

import requests
import time
from openpyxl import load_workbook
from collections import defaultdict

# World Bank API base URL
WB_API_BASE = "https://api.worldbank.org/v2/country/{country}/indicator/{indicator}"

# Poverty indicators to fetch
INDICATORS = {
    "SI.POV.DDAY": "poverty_headcount_3_dollars_pct",  # Poverty at $3.00/day
    "SI.POV.LMIC": "poverty_headcount_4_dollars_pct",  # Poverty at $4.20/day
    "SI.POV.UMIC": "poverty_headcount_8_dollars_pct",  # Poverty at $8.30/day
    "SI.POV.GINI": "gini_index",  # Gini index
}

def fetch_indicator_data(country_id, indicator_id):
    """Fetch indicator data from World Bank API for a country."""
    url = WB_API_BASE.format(country=country_id, indicator=indicator_id)
    params = {
        "format": "json",
        "per_page": 100,
        "date": "1960:2024"
    }

    try:
        response = requests.get(url, params=params, timeout=30)
        response.raise_for_status()
        data = response.json()

        if len(data) < 2 or data[1] is None:
            return {}

        # Parse the data into {year: value} format
        result = {}
        for item in data[1]:
            year = int(item["date"])
            value = item["value"]
            if value is not None:
                result[year] = value

        return result
    except Exception as e:
        print(f"  Error fetching {indicator_id} for {country_id}: {e}")
        return {}

def main():
    # Load the Excel file
    input_file = "/home/dan/world_bank_mcp/world_bank_gdp_data_billions.xlsx"
    print(f"Loading {input_file}...")

    wb = load_workbook(input_file)
    ws = wb.active

    # Get headers and find the column positions
    headers = [cell.value for cell in ws[1]]
    country_col = headers.index("country_id") + 1
    year_col = headers.index("year") + 1

    # Get unique countries
    countries = set()
    for row in ws.iter_rows(min_row=2, max_col=country_col, values_only=True):
        if row[0]:
            countries.add(row[0])

    countries = sorted(countries)
    print(f"Found {len(countries)} unique countries")

    # Add new column headers
    start_col = len(headers) + 1
    col_mapping = {}
    for i, (indicator_id, col_name) in enumerate(INDICATORS.items()):
        col_idx = start_col + i
        ws.cell(row=1, column=col_idx, value=col_name)
        col_mapping[indicator_id] = col_idx

    print(f"Added {len(INDICATORS)} new columns starting at column {start_col}")

    # Fetch and store all poverty data
    all_data = {}  # {country_id: {indicator_id: {year: value}}}

    total_countries = len(countries)
    for idx, country_id in enumerate(countries):
        print(f"[{idx+1}/{total_countries}] Fetching data for {country_id}...")
        all_data[country_id] = {}

        for indicator_id in INDICATORS.keys():
            data = fetch_indicator_data(country_id, indicator_id)
            all_data[country_id][indicator_id] = data
            time.sleep(0.1)  # Be nice to the API

    print("\nWriting data to Excel...")

    # Now write the data to the Excel file
    rows_updated = 0
    for row_idx in range(2, ws.max_row + 1):
        country_id = ws.cell(row=row_idx, column=country_col).value
        year = ws.cell(row=row_idx, column=year_col).value

        if country_id and year:
            year = int(year)
            country_data = all_data.get(country_id, {})

            for indicator_id, col_idx in col_mapping.items():
                indicator_data = country_data.get(indicator_id, {})
                value = indicator_data.get(year)
                if value is not None:
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    rows_updated += 1

    print(f"Updated {rows_updated} cells")

    # Save the file
    output_file = "/home/dan/world_bank_mcp/world_bank_gdp_data_with_poverty.xlsx"
    print(f"Saving to {output_file}...")
    wb.save(output_file)
    print("Done!")

if __name__ == "__main__":
    main()
